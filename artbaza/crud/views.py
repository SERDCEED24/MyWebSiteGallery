import os

from django.core.exceptions import ObjectDoesNotExist
from django.shortcuts import get_object_or_404, redirect, render
from django.apps import apps
from django.http import HttpResponseNotAllowed, HttpResponseBadRequest, HttpResponse, FileResponse
import json
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image as PILImage
from io import BytesIO

from artbaza import settings
from .models import Author, Genre, Material, Technique, WorkStatus
from .forms import MODEL_FORMS, MODEL_TITLES_HEADERS
from .constants import MODEL_TABLE_HEAD
from .models import Artwork
from django.db.models import Q
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.core.paginator import Paginator
import numpy as np
'''
def index(request):
    authors = Author.objects.all()
    return render(request, 'base_index.html', {'authors': authors})
'''
def format_number(num):
    if num.is_integer():
        return f'{num:.0f}'
    return f'{num:.2f}'

@login_required
def index(request, model_name):
    model = apps.get_model(app_label='crud', model_name=model_name)
    if not model:
        return HttpResponseBadRequest("Модель не найдена")
    if model_name == 'Artwork':
        fields = ['id', 'image', 'title', 'creation_year', 'author', 'status', 'current_price']
    else:
        fields = [field.name for field in model._meta.fields]

    query = request.GET.get("q", "").strip()

    records = model.objects.all()
    if query:
        match model_name:
            case 'Author':
                records = records.filter(Q(last_name__icontains=query) | Q(id__icontains=str(query)))
            case 'Genre':
                records = records.filter(Q(genre_name__icontains=query) | Q(id__icontains=str(query)))
            case 'Material':
                records = records.filter(Q(material_name__icontains=query) | Q(id__icontains=str(query)))
            case 'Technique':
                records = records.filter(Q(technique_name__icontains=query) | Q(id__icontains=str(query)))
            case 'WorkStatus':
                records = records.filter(Q(status_name__icontains=query) | Q(id__icontains=str(query)))
            case 'Artwork':
                records = records.filter(Q(title__icontains=query) | Q(id__icontains=str(query)))

    sort_field = request.GET.get('sort', '')
    direction = request.GET.get('direction', 'asc')

    if sort_field and sort_field in fields:
        if direction == 'desc':
            records = records.order_by(f'-{sort_field}')
        else:
            records = records.order_by(sort_field)

    field_label = []
    for i in range(len(fields)):
        field_label.append([fields[i], MODEL_TABLE_HEAD.get(model_name)[i]])

    # Пагинация
    per_page = request.GET.get('per_page', 6)  # По умолчанию 6
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 6
    # Ограничиваем возможные значения
    if per_page not in [3, 6, 12, 18]:
        per_page = 6

    paginator = Paginator(records, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Статистика
    stats_text = ''
    if model_name == 'Artwork':
        # Получаем все произведения
        artworks = Artwork.objects.all()

        # Общее количество
        total_count = artworks.count()

        # Если нет произведений, возвращаем нули
        if total_count == 0:
            stats_text = 'Произведений нет. Вычисление статистики невозможно!'
        else:
            # Получаем все цены
            prices = []
            for artwork in artworks:
                if artwork.current_price:
                    prices.append(artwork.current_price)

            # Расчет статистики
            total_price = sum(prices)
            avg_price = total_price / total_count
            median_price = np.median(prices)
            min_price = min(prices)
            max_price = max(prices)
            std_price = np.std(prices)  # Стандартное отклонение как разброс

            # Формируем текст для шаблона
            stats_text = (
                f'Количество произведений: {total_count}, '
                f'Стоимость коллекции: {format_number(total_price)}, '
                f'Минимальная цена: {format_number(min_price)}, '
                f'Средняя цена: {format_number(avg_price)}, '
                f'Медианная цена: {format_number(median_price)}, '
                f'Максимальная цена: {format_number(max_price)}, '
                f'Среднее отклонение от средней цены: {format_number(std_price)}'
            )

    return render(request, 'index.html', {
            'model_name': model_name,
            'fields': fields,
            'records': records,
            'sort_field': sort_field,
            'table_head': MODEL_TABLE_HEAD.get(model_name),
            'field_label': field_label,
            'direction': direction,
            'stats_text': stats_text,
            'page_obj': page_obj,
            'per_page': per_page
    })

@login_required
def delete_record(request, model_name, pk):
    if request.method == "POST":
        model = apps.get_model(app_label='crud', model_name=model_name)
        if not model:
            return JsonResponse({"error": "Модель не найдена"}, status=400)

        record = get_object_or_404(model, pk=pk)
        record.delete()
        return JsonResponse({"success": True, "record_id": pk})

    return JsonResponse({"error": "Метод не разрешён"}, status=405)

def delete_treenode(request, model_name, pk):
    try:
        model_map = {
            'Author': Author,
            'Genre': Genre,
            'Material': Material,
            'Technique': Technique,
            'WorkStatus': WorkStatus,
            'Artwork': Artwork
        }

        model = model_map.get(model_name)
        if not model:
            return JsonResponse({"success": False, "error": "Модель не найдена"}, status=400)

        obj = model.objects.get(id=pk)
        obj.delete()

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=400)

@login_required
def form(request, model_name, pk=None):
    model = apps.get_model(app_label='crud', model_name=model_name)
    if not model:
        return HttpResponseBadRequest("Модель не найдена")

    form_class = MODEL_FORMS.get(model_name)
    if not form_class:
        return HttpResponseBadRequest("Форма для этой модели не найдена")

    record = get_object_or_404(model, id=pk) if pk else None

    if request.method == "GET":
        referer = request.META.get('HTTP_REFERER', None)
        if referer:
            request.session['referer'] = referer

    referer = request.session.get('referer', None)
    if not referer:
        return redirect(f'/crud/{model_name}/')

    if request.method == "POST":
        model_form = form_class(request.POST, request.FILES, instance=record)
        if model_form.is_valid():
            model_form.save()
            return redirect(referer)
    else:
        model_form = form_class(instance=record)

    return render(request, 'form.html', {
        'form': model_form,
        'model_title': MODEL_TITLES_HEADERS[model_name][0],
        'model_header': MODEL_TITLES_HEADERS[model_name][1],
        'model_name' : model_name
    })

@login_required
def artwork_detailed(request, pk):
    artwork = Artwork.objects.get(pk=pk)
    return render(request, 'artwork_detailed.html', {'artwork': artwork})

def register(request):
    if request.method == 'POST':
        form_ = UserCreationForm(request.POST)
        if form_.is_valid():
            user = form_.save()
            login(request, user)
            return redirect('/crud/Artwork/')
    else:
        form_ = UserCreationForm()
    return render(request, 'register.html', {'form': form_})

def tree_view(request):
    context = {
        'authors': Author.objects.all(),
        'genres': Genre.objects.all(),
        'materials': Material.objects.all(),
        'techniques': Technique.objects.all(),
        'statuses': WorkStatus.objects.all(),
        'artworks': Artwork.objects.all()
    }
    return render(request, 'tree_view.html', context)


@login_required
def delete_multiple(request, model_name):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ids = data.get("ids", [])
            if not ids:
                return JsonResponse({"success": False, "error": "Нет выбранных записей"})
            model = apps.get_model(app_label='crud', model_name=model_name)
            if not model:
                return HttpResponseBadRequest("Модель не найдена")
            model.objects.filter(id__in=ids).delete()
            return JsonResponse({"success": True})
        except (ObjectDoesNotExist, json.JSONDecodeError) as e:
            return JsonResponse({"success": False, "error": str(e)})
    return JsonResponse({"success": False, "error": "Неверный метод запроса"})

@login_required
def export_to_excel(request):
    # Создаем новый Excel-файл
    wb = Workbook()
    # Удаляем дефолтный лист
    wb.remove(wb.active)

    # Список моделей и их полей
    models = ['Author', 'Genre', 'Material', 'Technique', 'WorkStatus', 'Artwork']
    model_fields = {
        'Author': ['id', 'last_name', 'first_name', 'middle_name', 'description', 'year_of_birth', 'year_of_death'],
        'Genre': ['id', 'genre_name'],
        'Material': ['id', 'material_name'],
        'Technique': ['id', 'technique_name'],
        'WorkStatus': ['id', 'status_name'],
        'Artwork': ['id', 'image', 'title', 'creation_year', 'author', 'length', 'width', 'height', 'genre', 'material', 'technique', 'status', 'purchase_year', 'purchase_price', 'current_price', 'price_without_frame', 'price_with_frame']
    }

    # Для каждой модели создаем лист
    for model_name in models:
        model = apps.get_model(app_label='crud', model_name=model_name)
        fields = model_fields[model_name]
        # Создаем лист с названием модели
        ws = wb.create_sheet(title=model_name)

        # Записываем заголовки
        for col, field in enumerate(fields, 1):
            ws.cell(row=1, column=col).value = field

        # Получаем все записи модели
        records = model.objects.all()

        # Заполняем данные
        for row, record in enumerate(records, 2):
            for col, field in enumerate(fields, 1):
                if field == 'image' and model_name in ['Author', 'Artwork'] and record.image:
                    try:
                        img_path = record.image.path
                        img = PILImage.open(img_path)
                        # Добавляем изображение без сжатия
                        img_buffer = BytesIO()
                        img.save(img_buffer, format='PNG')
                        img_buffer.seek(0)
                        openpyxl_img = OpenpyxlImage(img_buffer)
                        cell_address = ws.cell(row=row, column=col).coordinate
                        ws.add_image(openpyxl_img, cell_address)

                        # Настраиваем размер строки и столбца
                        ws.row_dimensions[row].height = img.height * 0.75  # Высота строки в пунктах
                        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = img.width / 7  # Примерная ширина столбца
                    except Exception:
                        ws.cell(row=row, column=col).value = "No image"
                else:
                    # Получаем значение поля
                    value = getattr(record, field, None)
                    if value and field in ['author', 'genre', 'material', 'technique', 'status']:
                        value = str(value)  # Для foreign key берем строковое представление
                    ws.cell(row=row, column=col).value = value if value is not None else "None"

    # Сохраняем файл в буфер
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Возвращаем файл для скачивания
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        content=output.getvalue()
    )
    response['Content-Disposition'] = 'attachment; filename="artwork_export.xlsx"'
    return response


@login_required
def download_help_pdf(request):
    # Путь к PDF-файлу в папке static
    pdf_path = os.path.join(settings.STATIC_ROOT, 'help', 'help.pdf')

    try:
        # Открываем файл и возвращаем его как ответ
        return FileResponse(
            open(pdf_path, 'rb'),
            as_attachment=True,
            filename='help.pdf'
        )
    except FileNotFoundError:
        return HttpResponse("Файл справки не найден", status=404)